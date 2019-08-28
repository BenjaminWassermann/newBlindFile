#!/usr/bin/env python3
# newBlindFile.py

# Generates a code sheet for blinded studies with n variables

# import statements
import openpyxl
from pathlib import Path
import pandas as pd
import xlwings as xl

# wrap function for potential importing
def main():

    # list which holds dictionary keys for ease of iteration?
    headers = ['Randomizer', 'Num', 'ID']

    # dictionary which holds column headers as keys
    # and lists of potential cell contents as values
    blindTable = {}
    for header in headers:
        blindTable[header] = []   
    

    # collect data of interest

    # collect desired total N
    n = int(input('How many animals need to be blinded?\n'))
    start = int(input('What number should I start with?\n'))

    # generate nums and add them to blindTable['Num']
    
    # loops through range of start through start + n
    blindTable['Num'].extend(range(start, start + n))

    # loop to collect number of surgical variables
    nextSurg = True
    surgeries = []
    
    while nextSurg:
        surg = input('Enter next surgery (blank if done):\n')

        # end loop by inputting ''
        if surg == '':
            nextSurg = False

        # else add a new surgery
        # add it to headers and the surgery to blindTable keys
        else:

            # if procedure isn't in headers, add it
            if 'Procedure' not in headers:
                headers.append('Procedure')

            # if surgery isn't already in  blindTable keys
            if surg not in blindTable.keys():
                num = input('How many animals should undergo %s?\n' % surg)
                blindTable[surg] = num
                surgeries.append(surg)

    # loop to collect number of dose variables
    nextDose = True
    doseNum = 0    

    while nextDose:
        dose = input('Enter next compound to be administered (blank if done):\n')

        # end loop by inputting ''
        if dose == '':
            nextDose = False

        # if blindTable doesn't have a doses entry
        # add one with an empty list
        elif 'Doses' not in blindTable.keys():
            headers.append('Dose')
            blindTable['Doses'] = []
            blindTable['Doses'].append(dose)
            doseNum += 1

        # else dose is a valid string and 'Doses' already in blind table
        # add dose to 'Doses'
        else:
            blindTable['Doses'].append(dose)
            doseNum += 1

    # data verification

    # make sure that the number of procedures adds up to the total number of animals
    surgSum = 0
    if surgeries:
        for surg in surgeries:
            surgSum += int(blindTable[surg])

        if surgSum != n:
            print('The desired n is %s, but there are %s surgeries requested...' % (n, surgSum))
            main()

        # make sure that procedure numbers divided by number of doses has no remainder
        for surg in surgeries:
            if doseNum != 0:
                if int(blindTable[surg]) % doseNum != 0:
                    print('%s %s procedures do not split evenly among %s doses...' % (blindTable[surg], surg, doseNum))
                    main()

    # else there are no surgeries, just split doses evenly
    elif doseNum != 0:

        if n % doseNum != 0:
            print('%s doses do not split evenly across %s animals...' % (doseNum, n))

    # generate excel spreadsheet

    # collect project name
    proj = input('What is the project code?\n')
    fileName = '%sBlind.xlsx' % proj
    path = Path().absolute()
    pathName = '%s\\%s' % (path, fileName)

    # open workbook and select sheet
    book1 = openpyxl.Workbook()
    sheet1 = book1.active

    # place headers across first row
    for i in range(1, len(headers) +1):
        sheet1.cell(row=1, column=i).value = headers[i-1]

    # fill Num column with contents of blindTable['Num']
    numCount = 2
    for num in blindTable['Num']:
        sheet1.cell(row=numCount, column=2).value = num
        numCount += 1

    # fill randomizer column with random values
    xlRand = '=rand()'
    for row in sheet1:
        if row[0].value != 'Randomizer':
            row[0].value = xlRand

    book1.save(pathName)
    book1.close()

    # save workbook using xlwings
    # gives excel a chance to calculate, allows Pandas to read formulas
    app = xl.App(visible=False)
    xlBook = app.books.open(fileName)
    xlBook.save()
    xlBook.close()
    app.kill()

    # use pandas to sort by randomizer
    df = pd.read_excel(fileName, sheet_name='Sheet')
    df.sort_values('Randomizer').to_excel(fileName, index=False)

    # fill in appropriate procedures and doses using openpyxl

    # load workbook
    book2 = openpyxl.load_workbook(pathName)
    sheet2 = book2.active
    book2.create_sheet('count')
    sheet3 = book2['count']

    # generate list containing the appropriate procedure names
    # the appropriate number of times
    fullPro = []
    fullDose = []    

    # runs if there are surgeries
    if surgeries:
        
        for surg in surgeries:
        
            fullPro.extend([surg] * int(blindTable[surg]))
            surgNum = int(blindTable[surg])

            # runs if there are surgeries and doses
            if 'Doses' in blindTable.keys():
            
                for dose in blindTable['Doses']:
            
                    dosePart = int(surgNum/doseNum)            
                    fullDose.extend([dose] * dosePart)

    # runs if there are doses but no surgeries
    elif 'Doses' in blindTable.keys():
            
        for dose in blindTable['Doses']:
            
            dosePart = int(n/doseNum)            
            fullDose.extend([dose] * dosePart)

    # count tracker
    cellCount = 0

    # iterate through rows
    for row in sheet2:
        
        # if a row is not the headers
        if row[0].value != 'Randomizer':

            # set procedure and dose entries

            # don't populate if there are no procedures
            if fullPro and fullDose:
                row[3].value = fullPro[cellCount]
                row[4].value = fullDose[cellCount]

            elif fullPro:
                row[3].value = fullPro[cellCount]

            elif fullDose:
                row[3].value = fullDose[cellCount]
                
            cellCount += 1

    # delete randomizer column
    sheet2.delete_cols(1)

    # save and close
    book2.save(pathName)
    book2.close()            

    # sort with pandas by num
    # use pandas to sort by Num
    df = pd.read_excel(fileName, sheet_name='Sheet1')
    df.sort_values('Num').to_excel(fileName, index=False)

    book3 = openpyxl.load_workbook(pathName)
    sheet3 = book3.active

    a = 1    

    if surgeries:

        for surg in surgeries:

            
            surgNum = int(blindTable[surg])

            if 'Doses' in blindTable.keys():

                for dose in blindTable['Doses']:

                    cell = 'f%s' % a
                    valueCell = 'g%s' %a
                
                    dosePart = int(surgNum/doseNum)      
                    sheet3[cell].value = '%s-%s' % (surg, dose)
                    sheet3[valueCell].value = dosePart

                    a += 1

            else:

                cell = 'f%s' % a
                valueCell = 'g%s' %a
                
                sheet3[cell].value = '%s' % surg
                sheet3[valueCell].value = surgNum

                a += 1


    elif 'Doses' in blindTable.keys():

        for dose in blindTable['Doses']:

            cell = 'f%s' % a
            valueCell = 'g%s' %a

            dosePart = int(n/doseNum)
            sheet3[cell].value = '%s' % dose
            sheet3[valueCell].value = dosePart

            a += 1
                

    # save and close
    book3.save(pathName)
    book3.close()

    print('%s updated!' % fileName)
    
# statement runs the main method if file is accessed but not when imported
if __name__ == '__main__':
    main()
